"""
Console entrypoint for the `gemini_code_review` command line tool.
"""

from __future__ import annotations

import argparse
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from .prompting import run_code_review  # type: ignore
except Exception:  # pragma: no cover - we handle absence later
    run_code_review = None  # type: ignore


def ensure_repomix_available() -> bool:
    """Return True if we can invoke repomix directly, else False."""
    try:
        subprocess.run(
            ["repomix", "--version"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            check=True,
        )
        return True
    except (subprocess.CalledProcessError, FileNotFoundError):
        return False


def dump_repo_to_xml(repo_path: Path, keep: bool = False) -> str:
    """Run repomix to dump repo to a temporary XML file and return its content."""
    if not repo_path.is_dir():
        raise ValueError(f"Provided path is not a directory: {repo_path}")

    with tempfile.NamedTemporaryFile(prefix="repomix_", suffix=".xml", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    cmd = [
        "repomix",
        "--quiet",
        "--style",
        "xml",
        "--output",
        str(tmp_path),
        str(repo_path),
    ]

    try:
        subprocess.run(cmd, check=True)
    except FileNotFoundError as e:
        if tmp_path.exists() and not keep:
            tmp_path.unlink(missing_ok=True)
        raise RuntimeError(
            "`repomix` not found. Please install repomix and ensure it is in your PATH."
        ) from e
    except subprocess.CalledProcessError as e:
        if tmp_path.exists() and not keep:
            tmp_path.unlink(missing_ok=True)
        raise RuntimeError(f"repomix failed (exit {e.returncode}).") from e

    try:
        content = tmp_path.read_text(encoding="utf-8")
    finally:
        if not keep:
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                pass
        else:
            print(f"[info] Kept temp XML at: {tmp_path}", file=sys.stderr)
    return content


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gemini Code Review CLI")
    parser.add_argument(
        "--path",
        dest="path",
        default=".",
        help="Repository path (default: current directory)",
    )
    parser.add_argument(
        "--keep",
        dest="keep",
        action="store_true",
        help="Keep the generated temporary XML file",
    )
    parser.add_argument(
        "--issues",
        dest="issues",
        type=int,
        default=None,
        help="Number of issues to request (default 10 if omitted)",
    )
    parser.add_argument(
        "--instructions",
        dest="instructions",
        type=str,
        default=None,
        help="Additional free-form user instructions for the model",
    )
    parser.add_argument(
        "--non-interactive",
        dest="non_interactive",
        action="store_true",
        help="Don't prompt; rely solely on flags (useful for scripts)",
    )
    return parser.parse_args(argv)


def prompt_for_int(prompt: str, default: int) -> int:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
    except EOFError:
        return default
    if not raw:
        return default
    try:
        val = int(raw)
        if val <= 0:
            raise ValueError
        return val
    except ValueError:
        print("Invalid number, using default.", file=sys.stderr)
        return default


def prompt_for_optional_text(prompt: str) -> Optional[str]:
    try:
        raw = input(f"{prompt} (leave blank for none): ").strip()
    except EOFError:
        return None
    return raw or None


def main(argv: list[str] | None = None) -> int:
    ns = parse_args(argv or sys.argv[1:])
    repo_path = Path(ns.path).resolve()

    if not ensure_repomix_available():
        print(
            "ERROR: repomix is not installed or not found in PATH. Please install repomix and try again.",
            file=sys.stderr,
        )
        return 2

    issues = ns.issues if ns.issues is not None else 10
    instructions = ns.instructions

    if not ns.non_interactive:
        if ns.issues is None:
            issues = prompt_for_int("How many issues would you like surfaced?", 10)
        if ns.instructions is None:
            instructions = prompt_for_optional_text("Any additional user instructions?")

    try:
        xml = dump_repo_to_xml(repo_path, keep=ns.keep)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if run_code_review is None:
        print(
            "ERROR: prompting module not available (check GOOGLE_API_KEY or dependencies).",
            file=sys.stderr,
        )
        return 3

    # Token count using tiktoken if available; fallback to simple heuristic
    token_count_method = "approx-heuristic"
    token_count = max(1, (len(xml) + 3) // 4)
    try:
        import tiktoken  # type: ignore

        try:
            enc = tiktoken.get_encoding("o200k_base")
        except Exception:
            # Fallback to a widely available base if o200k_base is not present
            enc = tiktoken.get_encoding("cl100k_base")
        token_count = len(enc.encode(xml))
        token_count_method = "tiktoken"
    except Exception:
        pass

    print(
        f"[info] Repository token count ({token_count_method}): {token_count:,}",
        file=sys.stderr,
    )

    print(f"[info] Requesting up to {issues} issues from model...", file=sys.stderr)

    try:
        review_response = run_code_review(
            codebase_xml=xml,
            n_issues_to_surface=issues,
            user_instructions=instructions,
        )
    except Exception as exc:
        print(f"ERROR: model invocation failed: {exc}", file=sys.stderr)
        return 4

    try:
        import pandas as pd

        rows = []
        for issue in review_response.issues:
            # Programmatically generated Markdown summary for easy copy/paste
            copy_paste_md = (
                f"### [{issue.category}] {issue.title}\n"
                f"- Severity: {issue.severity}\n"
                f"- Location: {issue.location}\n"
                f"- Effort: {issue.estimated_effort}\n"
                f"- Rationale: {issue.rationale}\n"
                f"- Implementation plan: {issue.implementation_plan}\n\n"
                f"{issue.detailed_description}"
            )

            rows.append(
                {
                    "category": issue.category,
                    "title": issue.title,
                    "severity": issue.severity,
                    "location": issue.location,
                    "estimated_effort": issue.estimated_effort,
                    "rationale": issue.rationale,
                    "implementation_plan": issue.implementation_plan,
                    "detailed_description": issue.detailed_description,
                    "copy_paste": copy_paste_md,
                }
            )
        desired_columns = [
            "category",
            "title",
            "severity",
            "location",
            "estimated_effort",
            "rationale",
            "detailed_description",
            "implementation_plan",
            "copy_paste",
        ]
        df = pd.DataFrame(rows)
        df = df.reindex(columns=desired_columns)

        # Timestamped output filename as requested: gemini_code_reivew_[mm]-[dd]-[yyyy]_[hh]-[mm].xlsx
        now = datetime.now()
        timestamp = now.strftime("%m-%d-%Y_%H-%M")
        out_filename = f"gemini_code_reivew_{timestamp}.xlsx"
        out_path = Path.cwd() / out_filename

        from openpyxl.styles import Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            sheet_name = "Review"
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            column_widths = {
                "category": 15,
                "title": 40,
                "severity": 10,
                "location": 25,
                "estimated_effort": 15,
                "rationale": 40,
                "implementation_plan": 60,
                "detailed_description": 60,
                "copy_paste": 20,
            }
            for col_index, column_name in enumerate(desired_columns, start=1):
                letter = get_column_letter(col_index)
                width = column_widths.get(column_name)
                if width is not None:
                    worksheet.column_dimensions[letter].width = width

            # Freeze header row and the first two columns (category, title)
            worksheet.freeze_panes = "C2"

            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    cell.alignment = wrap_alignment

            # Ensure the copy_paste column does not wrap
            try:
                copy_col_index = desired_columns.index("copy_paste") + 1
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=copy_col_index)
                    cell.alignment = Alignment(wrap_text=False, vertical="top")
            except ValueError:
                pass

            fill_for_value = {
                "Low": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
                "Medium": PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
                "High": PatternFill(start_color="FFF8CBAD", end_color="FFF8CBAD", fill_type="solid"),
                "Critical": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
                "Very High": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
            }

            severity_col_index = desired_columns.index("severity") + 1
            effort_col_index = desired_columns.index("estimated_effort") + 1

            for row_idx in range(2, worksheet.max_row + 1):
                sev_cell = worksheet.cell(row=row_idx, column=severity_col_index)
                sev_value = str(sev_cell.value).strip() if sev_cell.value is not None else ""
                if sev_value in fill_for_value:
                    sev_cell.fill = fill_for_value[sev_value]

                eff_cell = worksheet.cell(row=row_idx, column=effort_col_index)
                eff_value = str(eff_cell.value).strip() if eff_cell.value is not None else ""
                if eff_value in fill_for_value:
                    eff_cell.fill = fill_for_value[eff_value]

            # ABOUT sheet with run metadata
            about_sheet_name = "ABOUT"
            about_rows = [
                {"Field": "Generated At", "Value": now.strftime("%Y-%m-%d %H:%M")},
                {"Field": "Repository Path", "Value": str(repo_path)},
                {"Field": "Output File", "Value": out_filename},
                {"Field": "Token Count", "Value": token_count},
                {"Field": "Token Count Method", "Value": token_count_method},
                {"Field": "Issues Requested", "Value": issues},
                {"Field": "Issues Returned", "Value": len(review_response.issues)},
                {"Field": "User Instructions", "Value": instructions or ""},
                {"Field": "Model", "Value": "gemini-2.5-pro"},
                {
                    "Field": "Notes",
                    "Value": (
                        "Token count is an approximation based on XML size (\u2248 1 token per 4 chars)."
                    ),
                },
            ]
            df_about = pd.DataFrame(about_rows, columns=["Field", "Value"])
            df_about.to_excel(writer, index=False, sheet_name=about_sheet_name)

            about_ws = writer.sheets[about_sheet_name]
            # Set reasonable column widths
            about_ws.column_dimensions[get_column_letter(1)].width = 24
            about_ws.column_dimensions[get_column_letter(2)].width = 90
            # Wrap text and top-align
            for row in about_ws.iter_rows(
                min_row=1,
                max_row=about_ws.max_row,
                min_col=1,
                max_col=about_ws.max_column,
            ):
                for cell in row:
                    cell.alignment = wrap_alignment

        print(f"[info] Wrote Excel report to {out_path}", file=sys.stderr)
    except Exception as exc:
        print(f"ERROR: failed to write Excel file: {exc}", file=sys.stderr)
        return 5

    try:
        import json
        from pydantic import BaseModel

        def model_to_dict(obj):
            if isinstance(obj, BaseModel):
                return obj.model_dump()
            if isinstance(obj, list):
                return [model_to_dict(o) for o in obj]
            return obj

        print(json.dumps(model_to_dict(review_response), indent=2))
    except Exception:
        pass

    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())


